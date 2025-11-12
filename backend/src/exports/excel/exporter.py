from decimal import Decimal
from io import BytesIO
from typing import Optional
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from ...models_quote import Quote, QuoteDay, QuoteLine
import math


def _days_count(q: Quote) -> int:
    """Calcule le nombre de jours à partir de la liste des jours ou de la différence de dates."""
    if q.days:
        return len(q.days)
    # sinon on tombe sur le diff de dates inclusif
    try:
        from datetime import date as dt_date
        d0 = q.start_date if isinstance(q.start_date, dt_date) else dt_date.fromisoformat(str(q.start_date))
        d1 = q.end_date   if isinstance(q.end_date,   dt_date) else dt_date.fromisoformat(str(q.end_date))
        return max(0, (d1 - d0).days + 1)
    except Exception:
        return 0


def compute_onspot(q: Quote) -> Decimal:
    """Calcule le montant Onspot avec minimum 3 jours par carte."""
    pax = q.pax or 0
    cards = max(1, math.ceil((pax or 1) / 6))
    trip_days = _days_count(q)
    effective_days = max(trip_days, 3)  # **minimum 3 jours par carte**
    auto_val = cards * 9 * effective_days
    return q.onspot_manual if q.onspot_manual is not None else Decimal(str(auto_val))


def _round2(val) -> float:
    """Round to 2 decimals."""
    if val is None:
        return 0.0
    try:
        return round(float(val), 2)
    except (ValueError, TypeError):
        return 0.0


def _parse_float(val) -> float:
    """Parse float value, return 0 if invalid."""
    if val is None:
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def _get_buff_pct(line: QuoteLine) -> float:
    """Extract buff_pct from line (raw_json or None)."""
    if line.raw_json and isinstance(line.raw_json, dict):
        buff = line.raw_json.get("buff_pct")
        if buff is not None:
            return _parse_float(buff)
    return 0.0


def _get_provider_url(line: QuoteLine) -> str:
    """Extract provider_service_url from line raw_json."""
    if line.raw_json and isinstance(line.raw_json, dict):
        # Check in fields first (from catalog, normalized via extract_excel_fields)
        fields = line.raw_json.get("fields", {})
        if isinstance(fields, dict) and fields.get("provider_service_url"):
            url = fields["provider_service_url"]
            if url:
                return str(url).strip()
        
        # Check in snapshot (from catalog full data)
        snapshot = line.raw_json.get("snapshot", {})
        if isinstance(snapshot, dict):
            snapshot_fields = snapshot.get("fields", {})
            if isinstance(snapshot_fields, dict) and snapshot_fields.get("provider_service_url"):
                url = snapshot_fields["provider_service_url"]
                if url:
                    return str(url).strip()
        
        # Fallback to direct in raw_json
        if line.raw_json.get("provider_service_url"):
            url = line.raw_json["provider_service_url"]
            if url:
                return str(url).strip()
    return ""


def _get_internal_note(line: QuoteLine) -> str:
    """Extract internal_note from line raw_json."""
    if line.raw_json and isinstance(line.raw_json, dict):
        note = line.raw_json.get("internal_note")
        if note:
            return str(note).strip()
    return ""


def _excel_service_name(line: QuoteLine) -> str:
    """Generate service name for Excel (matching frontend logic)."""
    cat = (line.category or "").lower()
    raw_json = line.raw_json or {}
    data = raw_json
    
    if cat == "flight":
        from_val = data.get("from", "?")
        to_val = data.get("to", "?")
        name = f"Flight {from_val}->{to_val}"
    elif cat == "train":
        from_val = data.get("from", "?")
        to_val = data.get("to", "?")
        name = f"Train {from_val}->{to_val}"
    elif cat == "ferry":
        from_val = data.get("from", "?")
        to_val = data.get("to", "?")
        name = f"Ferry {from_val}->{to_val}"
    elif cat in ("hotel", "new hotel"):
        name = data.get("hotel_name") or line.title or "Hotel"
    elif cat in ("activity", "new service"):
        name = data.get("title") or line.title or "Service"
    elif cat == "car rental":
        name = "Car Rental"
    elif cat == "cost":
        name = data.get("title") or line.title or "Cost"
    else:
        name = line.title or "—"
    
    # Clamp to 50 chars
    if len(name) > 50:
        name = name[:49] + "…"
    return name


def _effective_fx(line_fx: Optional[Decimal], quote_fx: Optional[Decimal], default: float = 1.0) -> float:
    """Get effective FX rate (line > quote > default)."""
    if line_fx is not None and float(line_fx) > 0:
        return float(line_fx)
    if quote_fx is not None and float(quote_fx) > 0:
        return float(quote_fx)
    return default


def _is_paid_category(category: Optional[str]) -> bool:
    """Check if category is paid (exclude Trip info, Internal)."""
    if not category:
        return False
    cat_lower = category.lower().strip()
    return cat_lower not in ("trip info", "internal", "internal info")


def _format_number(val: float) -> tuple[str, str]:
    """
    Return (formatted_value, number_format) for Excel.
    - If integer: format "0", display as integer (no decimals)
    - If decimal: format "0.00", display with 2 decimals
    - If zero: return empty string and format that hides zero
    """
    if val == 0.0:
        return ("", "0;;;")  # Format that shows blank for zero
    
    if val == int(val):
        return (str(int(val)), "0;;;")  # Integer format, no decimals, hide zero
    else:
        return (str(round(val, 2)), "0.00;;;")  # 2 decimals, hide zero


def build_xlsx_for_quote(db: Session, quote_id: int) -> BytesIO:
    """
    Build Excel export v2 for a quote.
    Single sheet with destinations, formulas, totals, and recap.
    """
    q = db.query(Quote).filter(Quote.id == quote_id).first()
    if not q:
        raise ValueError(f"Quote {quote_id} not found")
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Quote"
    
    # Set column widths
    ws.column_dimensions[get_column_letter(1)].width = 17    # Column A
    ws.column_dimensions[get_column_letter(2)].width = 12   # Column B
    ws.column_dimensions[get_column_letter(3)].width = 41.87 # Column C
    ws.column_dimensions[get_column_letter(4)].width = 20.8 # Column D
    ws.column_dimensions[get_column_letter(5)].width = 20.8 # Column E
    ws.column_dimensions[get_column_letter(6)].width = 11.3 # Column F
    ws.column_dimensions[get_column_letter(7)].width = 21   # Column G
    ws.column_dimensions[get_column_letter(8)].width = 10   # Column H
    ws.column_dimensions[get_column_letter(9)].width = 20   # Column I
    
    # Default FX (fallback)
    default_fx = 1.0
    if q.fx_rate:
        default_fx = float(q.fx_rate)
    
    # Compute Onspot and Hassle
    onspot_total = compute_onspot(q)
    onspot_value = float(onspot_total) if onspot_total else 0.0
    
    hassle_auto = (q.pax or 0) * 150
    hassle_value = float(q.hassle_manual) if q.hassle_manual is not None else hassle_auto
    
    # Collect all paid service rows
    service_rows = []
    current_dest = None
    
    # Sort days by position
    days = sorted(q.days, key=lambda d: d.position or 0)
    
    for day in days:
        # Sort lines by position
        lines = sorted(day.lines, key=lambda l: l.position or 0)
        
        for line in lines:
            if not _is_paid_category(line.category):
                continue
            
            # Destination (only on first service of each destination)
            dest = day.destination or ""
            show_dest = (dest != current_dest)
            if show_dest:
                current_dest = dest
            
            # Service name
            service_name = _excel_service_name(line)
            
            # Purchase EUR raw value
            purchase_eur_raw = float(line.achat_eur) if line.achat_eur else 0.0
            buff_pct = _get_buff_pct(line)
            has_buff = buff_pct and buff_pct > 0
            
            # FX rate
            fx = _effective_fx(line.fx_rate, q.fx_rate, default_fx)
            
            # Sell USD
            sell_usd = float(line.vente_usd) if line.vente_usd else 0.0
            
            # Supplier
            supplier = line.supplier_name or ""
            
            # Internal note
            internal_note = _get_internal_note(line)
            
            # Provider URL
            provider_url = _get_provider_url(line)
            
            service_rows.append({
                "dest": dest if show_dest else "",
                "service_name": service_name,
                "purchase_eur_raw": purchase_eur_raw,
                "buff_pct": buff_pct if has_buff else None,
                "has_buff": has_buff,
                "fx": fx,
                "sell_usd": sell_usd,
                "supplier": supplier,
                "internal_note": internal_note,
                "provider_url": provider_url,
            })
    
    # === HEADERS (Row 2) ===
    headers = ["Destination", "Service", "Prix d'achat €", "Prix d'achat $", "Prix de vente", "Supplier", "Note", "Provider URL"]
    for col_idx, header in enumerate(headers, start=2):  # Start at column B
        cell = ws.cell(row=2, column=col_idx)
        cell.value = header
        cell.alignment = Alignment(horizontal="center")
    
    # === SPECIAL ROWS ===
    # Row 3: Hassle (F3)
    ws.cell(row=3, column=6).value = hassle_value
    hassle_fmt = "0;;;" if hassle_value == int(hassle_value) else "0.00;;;"
    ws.cell(row=3, column=6).number_format = hassle_fmt
    
    # Row 4: Onspot (E4)
    ws.cell(row=4, column=5).value = onspot_value
    onspot_fmt = "0;;;" if onspot_value == int(onspot_value) else "0.00;;;"
    ws.cell(row=4, column=5).number_format = onspot_fmt
    
    # === SERVICE ROWS (starting at row 5) ===
    first_service_row = 5
    last_service_row = first_service_row + len(service_rows) - 1 if service_rows else first_service_row - 1
    
    for idx, row_data in enumerate(service_rows):
        excel_row = first_service_row + idx
        
        # Column B: Destination
        if row_data["dest"]:
            ws.cell(row=excel_row, column=2).value = row_data["dest"]
        
        # Column C: Service
        ws.cell(row=excel_row, column=3).value = row_data["service_name"]
        
        # Column D: Purchase EUR (with formula if buff_pct is present)
        if row_data["purchase_eur_raw"] > 0:
            purchase_eur_val = row_data["purchase_eur_raw"]
            if row_data["has_buff"]:
                # Formula with raw value visible: ={purchase_eur_raw}*(1+{buff_pct}/100)
                buff_pct_val = row_data["buff_pct"]
                ws.cell(row=excel_row, column=4).value = f"={purchase_eur_val}*(1+{buff_pct_val}/100)"
                # Calculate expected result to determine format
                expected_result = purchase_eur_val * (1 + buff_pct_val / 100)
                eur_fmt = "0;;;" if expected_result == int(expected_result) else "0.00;;;"
            else:
                # Direct value
                ws.cell(row=excel_row, column=4).value = purchase_eur_val
                eur_fmt = "0;;;" if purchase_eur_val == int(purchase_eur_val) else "0.00;;;"
            ws.cell(row=excel_row, column=4).number_format = eur_fmt
        else:
            ws.cell(row=excel_row, column=4).number_format = "0;;;"
        
        # Column E: Purchase USD (formula: D{row} / FX)
        if row_data["purchase_eur_raw"] > 0 and row_data["fx"] > 0:
            # Formula: D{row} / FX
            fx_val = row_data["fx"]
            ws.cell(row=excel_row, column=5).value = f"=D{excel_row}/{fx_val}"
            # Calculate expected result to determine format
            purchase_eur_for_usd = row_data["purchase_eur_raw"]
            if row_data["has_buff"]:
                purchase_eur_for_usd = purchase_eur_for_usd * (1 + row_data["buff_pct"] / 100)
            expected_usd = purchase_eur_for_usd / fx_val
            usd_fmt = "0;;;" if expected_usd == int(expected_usd) else "0.00;;;"
            ws.cell(row=excel_row, column=5).number_format = usd_fmt
        else:
            ws.cell(row=excel_row, column=5).number_format = "0;;;"
        
        # Column F: Sell USD
        sell_val, sell_fmt = _format_number(row_data["sell_usd"])
        if sell_val:
            ws.cell(row=excel_row, column=6).value = float(sell_val)
        ws.cell(row=excel_row, column=6).number_format = sell_fmt
        
        # Column G: Supplier
        if row_data["supplier"]:
            ws.cell(row=excel_row, column=7).value = row_data["supplier"]
        
        # Column H: Internal note
        if row_data["internal_note"]:
            ws.cell(row=excel_row, column=8).value = row_data["internal_note"]
        
        # Column I: Provider URL (as clickable hyperlink)
        if row_data["provider_url"]:
            url = row_data["provider_url"].strip()
            # Ensure URL has protocol
            if url and not url.startswith(("http://", "https://")):
                url = "https://" + url
            cell = ws.cell(row=excel_row, column=9)
            cell.value = url
            cell.hyperlink = url
            cell.font = Font(underline="single", color="0563C1")
    
    # === BLANK ROW ===
    blank_row = last_service_row + 1
    
    # === TOTAL ROW ===
    total_row = blank_row + 1
    ws.cell(row=total_row, column=2).value = "Total"
    ws.cell(row=total_row, column=2).font = Font(bold=True)
    
    # Column E: SUM of all service E values + E4 (Onspot)
    if last_service_row >= first_service_row:
        ws.cell(row=total_row, column=5).value = f"=SUM(E{first_service_row}:E{last_service_row})+E4"
    else:
        ws.cell(row=total_row, column=5).value = "=E4"
    ws.cell(row=total_row, column=5).number_format = "0.##;;;"  # Show decimals only if needed
    
    # Column F: SUM of all service F values + F3 (Hassle)
    if last_service_row >= first_service_row:
        ws.cell(row=total_row, column=6).value = f"=SUM(F{first_service_row}:F{last_service_row})+F3"
    else:
        ws.cell(row=total_row, column=6).value = "=F3"
    ws.cell(row=total_row, column=6).number_format = "0.##;;;"  # Show decimals only if needed
    
    # === GRAND TOTAL ROW ===
    grand_total_row = total_row + 1
    ws.cell(row=grand_total_row, column=2).value = "Grand total"
    ws.cell(row=grand_total_row, column=2).font = Font(bold=True)
    
    # Column F: (Total purchases $ incl. Onspot) + (Total sales incl. Hassle)
    ws.cell(row=grand_total_row, column=6).value = f"=E{total_row}+F{total_row}"
    ws.cell(row=grand_total_row, column=6).number_format = "0.##;;;"  # Show decimals only if needed
    
    # === BORDERS ===
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Apply borders from row 2 to grand_total_row
    for row in range(2, grand_total_row + 1):
        for col in range(2, 10):  # Columns B to I
            ws.cell(row=row, column=col).border = thin_border
    
    # === RECAP BLOCK (2 blank rows after grand total) ===
    recap_start_row = grand_total_row + 3
    
    # R1: Prix d'achat
    ws.cell(row=recap_start_row, column=3).value = "Prix d'achat"
    ws.cell(row=recap_start_row, column=4).value = f"=E{total_row}"
    ws.cell(row=recap_start_row, column=4).number_format = "0.##;;;"  # Show decimals only if needed
    
    # R2: Commission
    margin_pct = float(q.margin_pct) if q.margin_pct else 0.1627
    ws.cell(row=recap_start_row + 1, column=3).value = "Commission"
    ws.cell(row=recap_start_row + 1, column=4).value = f"=E{total_row}*{margin_pct}"
    ws.cell(row=recap_start_row + 1, column=4).number_format = "0.##;;;"  # Show decimals only if needed
    
    # R3: Prix de vente
    ws.cell(row=recap_start_row + 2, column=3).value = "Prix de vente"
    ws.cell(row=recap_start_row + 2, column=4).value = f"=F{total_row}"
    ws.cell(row=recap_start_row + 2, column=4).number_format = "0.##;;;"  # Show decimals only if needed
    
    # R4: Total
    ws.cell(row=recap_start_row + 3, column=3).value = "Total"
    ws.cell(row=recap_start_row + 3, column=3).font = Font(bold=True)
    ws.cell(row=recap_start_row + 3, column=4).value = f"=D{recap_start_row}+D{recap_start_row + 1}+D{recap_start_row + 2}"
    ws.cell(row=recap_start_row + 3, column=4).number_format = "0.##;;;"  # Show decimals only if needed
    
    # Set row height to 14.3 for all rows (from row 2 to last recap row)
    last_row = recap_start_row + 3
    for row_num in range(2, last_row + 1):
        ws.row_dimensions[row_num].height = 14.3
    
    # Save to BytesIO
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

